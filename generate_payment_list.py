
import sys
import openpyxl
import datetime
import traceback
import os
import zipfile

from openpyxl						import Workbook
from openpyxl						import load_workbook
from openpyxl.styles				import Color, PatternFill, Font, Border, Side, Alignment, Protection, colors
from copy 							import copy
from openpyxl.comments  			import Comment
from openpyxl.worksheet				import Worksheet
from openpyxl 						import *
from openpyxl.utils 				import get_column_letter
from datetime 						import timedelta

nme_lst = []
max_payment_lst = [
	{'yer': 2016, 'one_max_pmnt' : 30000, 'all_max_pmnt': 45000},
	{'yer': 2017, 'one_max_pmnt' : 33000, 'all_max_pmnt': 49500},
	{'yer': 2018, 'one_max_pmnt' : 36000, 'all_max_pmnt': 54000},
	{'yer': 2019, 'one_max_pmnt' : 40000, 'all_max_pmnt': 60000},
]


def elog(name, msg):
	
	"""
		creates log file with Args:
		name 	- string - name of log file
		msg 	- string - text of error message
		mem_loc - string - location in memory for saving log_file
	"""
	with open('{path}logs\\elog_{name}.txt'.format (path = path, name = name), "a") as log_file:
		log_file.write('\n{sep}'.format(sep = '--------------------------------------------------------------------'))
		log_file.write('\n{msg}'.format(msg = msg))
		log_file.write('\n{sep}'.format(sep = '--------------------------------------------------------------------'))


def openpyxl_border_range(ws, start_col, start_row, end_col, end_row):

	top_lft_corner_border = Border(top = Side(border_style='thin', color=colors.BLACK), left  = Side(border_style='thin', color=colors.BLACK),)
	top_rgt_corner_border = Border(top = Side(border_style='thin', color=colors.BLACK), right = Side(border_style='thin', color=colors.BLACK),)
	bot_lft_corner_border = Border(bottom = Side(border_style='thin', color=colors.BLACK), left  = Side(border_style='thin', color=colors.BLACK),) 
	bot_rgt_corner_border = Border(bottom = Side(border_style='thin', color=colors.BLACK), right = Side(border_style='thin', color=colors.BLACK),) 
	all_side_border = Border(top = Side(border_style='thin', color=colors.BLACK), left=Side(border_style='thin', color=colors.BLACK),
											bottom = Side(border_style='thin', color=colors.BLACK), right=Side(border_style='thin', color=colors.BLACK),)

	top_side_border = Border(top=Side(border_style='thin', color=colors.BLACK),)
	rgt_side_border = Border(right=Side(border_style='thin', color=colors.BLACK),)
	bot_side_border = Border(bottom=Side(border_style='thin', color=colors.BLACK),)
	lft_side_border = Border(left=Side(border_style='thin', color=colors.BLACK),)

	top_lft_corner_cell = [start_col, start_row]
	top_rgt_corner_cell = [end_col, start_row]
	bot_lft_corner_cell = [start_col, end_row]
	bot_rgt_corner_cell = [end_col, end_row]

	top_side_row = start_row
	bot_side_row = end_row
	lft_side_col = start_col
	rgt_side_col = end_col

	top_side_cell_start = [lft_side_col+1, top_side_row]
	top_side_cell_end = [rgt_side_col-1, top_side_row]
	rgt_side_cell_start = [rgt_side_col, top_side_row + 1]
	rgt_side_cell_end = [rgt_side_col, bot_side_row - 1]
	bot_side_cell_start = [lft_side_col+1, bot_side_row]
	bot_side_cell_end = [rgt_side_col-1, bot_side_row]
	lft_side_cell_start = [lft_side_col, top_side_row + 1]
	lft_side_cell_end = [lft_side_col, bot_side_row - 1]

	col = get_column_letter(top_lft_corner_cell[0])
	row=top_lft_corner_cell[1]
	ws['%s%s'%(col, row)].border = top_lft_corner_border

	col = get_column_letter(top_rgt_corner_cell[0])
	row=top_rgt_corner_cell[1]
	ws['%s%s'%(col, row)].border = top_rgt_corner_border

	col = get_column_letter(bot_lft_corner_cell[0])
	row=bot_lft_corner_cell[1]
	ws['%s%s'%(col, row)].border = bot_lft_corner_border

	col = get_column_letter(bot_rgt_corner_cell[0])
	row=bot_rgt_corner_cell[1]
	ws['%s%s'%(col, row)].border = bot_rgt_corner_border

	for col_idx in range(top_side_cell_start[0], top_side_cell_end[0] + 1):
		col = get_column_letter(col_idx)
		row=top_side_row
		ws['%s%s'%(col, row)].border = top_side_border

	for col_idx in range(bot_side_cell_start[0], bot_side_cell_end[0] + 1):
		col = get_column_letter(col_idx)
		row=bot_side_row
		ws['%s%s'%(col, row)].border = bot_side_border

	for row_idx in range(lft_side_cell_start[1], lft_side_cell_end[1] + 1):
		col = get_column_letter(lft_side_col)
		row=row_idx
		ws['%s%s'%(col, row)].border = lft_side_border

	for row_idx in range(rgt_side_cell_start[1], rgt_side_cell_end[1] + 1):
		col = get_column_letter(rgt_side_col)
		row=row_idx
		ws['%s%s'%(col, row)].border = rgt_side_border


def group_fill_color(ws, start_col, start_row, end_col, end_row, color):
	
	for col_index in range(start_col, end_col + 1):
		for row_index in range(start_row, end_row + 1):
			ws.cell(row=row_index, column=col_index).fill = PatternFill(
				start_color=color,
				end_color=color,
				fill_type='solid'
			)


def load_bfs(filename):
	
	try:
			
		list_pay = []
			
		se_rfr_str = 'SE Reference'.lower()
		cn_amt_str = 'Amount (converted)'.lower()
		op_nme_str = 'Local Opportunity Name'.lower()
		op_lde_str = 'Opportunity Leader'.lower()
		cls_dte_str = 'Close Date'.lower()
		sls_stg_str = 'Phase/Sales Stage'.lower()
				
		wb_bfs = load_workbook(filename)
		sheet_ranges = wb_bfs[u'{x}'.format(x = find_corect_list (wb_bfs))]#initiating source sheet
		
		#looking for max row number
		for row_index in range(1,1000):
			if sheet_ranges.cell(row=row_index, column=1).value is None:
			#if str(sheet_ranges.cell(row=row_index, column=1).value).lower() == 'MY_FSR_Leads'.lower():
				sht_max_row=row_index - 1
		
		for col_index in range(1,50):
			cur_cell = str(sheet_ranges.cell(row=1, column=col_index).value).lower()
			if se_rfr_str in cur_cell	:	se_rfr_col = col_index
			if cn_amt_str == cur_cell	:	cn_amt_col = col_index
			if op_nme_str in cur_cell	:	op_nme_col = col_index
			if op_lde_str in cur_cell	:	op_lde_col = col_index
			if cls_dte_str in cur_cell	:	cls_dte_col = col_index
			if sls_stg_str in cur_cell	:	sls_stg_col = col_index
					
		for row_index in range(2, sht_max_row + 1):
			#creating list_pay
			if   ((sheet_ranges.cell(row=row_index, column=op_nme_col).value is not None) and
				  (sheet_ranges.cell(row=row_index, column=se_rfr_col).value != '')) :
				#if ((datetime.datetime.date(sheet_ranges.cell(row=row_index, column=cls_dte_col).value) >= date_start_month) and
				# (datetime.datetime.date(sheet_ranges.cell(row=row_index, column=cls_dte_col).value) <= date_end_month)):
				
				list_pay_line = {}
				
				list_pay_line['se_rfr'] = str(sheet_ranges.cell(row=row_index, column=se_rfr_col).value)
				list_pay_line['cn_amt'] = float(sheet_ranges.cell(row=row_index, column=cn_amt_col).value)
				list_pay_line['op_nme'] = str(sheet_ranges.cell(row=row_index, column=op_nme_col).value)
				list_pay_line['op_lde'] = str(sheet_ranges.cell(row=row_index, column=op_lde_col).value)
				
				list_pay.append(list_pay_line)

		return list_pay

	except Exception as e:
		elog('load_bfs','{d}_{type}\n{trb1}\n{trb2} '.format(
			d=datetime.datetime.now(),
			type=str(type(e)),
			trb1=traceback.format_tb(sys.exc_info()[2])[0],
			trb2=str(sys.exc_info()[1])
			)
		)
		return []


def list_sep_by_nme(list_pay):
	
	try:
		
		list_sep_1 = []
		list_sep_2 = []
		list_final_by_nme = []
		list_nms= []
		
		for line in list_pay:
			#print (line)
			new_line_1					 = line
			new_line_1['op_nme']		 = str(new_line_1['op_nme']).lower()
			new_line_1['op_nme']		 = new_line_1['op_nme'].split(',')		
			list_sep_1.append(new_line_1)
		
		for line in list_sep_1:
			lnth = len(line['op_nme'])
			
			for name in line['op_nme']:
				
				if 'FSR'.lower() in name.lower()	:	name = name.replace('fsr ','')
				elif 'SC'.lower() in name.lower()	:	name = name.replace('sc ','')
				
				new_line_2 = {}
				new_line_2['op_nme'] = str(name).strip(' ')
				
				new_line_2['se_rfr'] = line['se_rfr']
				new_line_2['cn_amt']= line['cn_amt']
				new_line_2['op_lde'] = line['op_lde'].strip(' ')

				
				if lnth == 1:	
					if float(line['cn_amt'])*0.01 < max_payment_lne['one_max_pmnt']:
						new_line_2['pmt_cnt'] = float(line['cn_amt'])*0.01
					else : 
						new_line_2['pmt_cnt'] = max_payment_lne['one_max_pmnt']
				
				else:
					if float(line['cn_amt'])*0.015 < max_payment_lne['all_max_pmnt']:
						new_line_2['pmt_cnt'] = (float(line['cn_amt'])*0.015)/lnth
					else:
						new_line_2['pmt_cnt'] = max_payment_lne['all_max_pmnt']/lnth
				
				list_sep_2.append(new_line_2)
		
		for line in list_sep_2:
			list_nms.append(line['op_nme'])
		
		list_nms = list(set(list_nms))
		list_nms.sort() 
		
		for element in list_nms:
		
			line_fnl = {}
			pmt_fnl= 0
			line_fnl['op_nme'] = str(element)
			nme_ms= element.split(' ')
			line_fnl['nme']= str(nme_ms[0])
			
			try:
				
				line_fnl['lst_nme']= str(nme_ms[1])
				line_fnl['op_lst'] = []
				line_fnl['ru_nme']= u''
			
			except:
				line_fnl['lst_nme'] = ''
				line_fnl['op_lst'] = []
				line_fnl['ru_nme'] = u''
				#print (nme_ms[0])
			
			for line_n in nme_lst:
				if line_n['bfs_nme'].lower() in str(element).lower() :
					line_fnl['ru_nme'] = line_n['ru_nme']
			
			if line_fnl['ru_nme']== '':
				print ('error: name "{name}" in is incorrect (not include in source list)'.format (name = str(element)))
				#print ('{WO} - error: name "{name}" in is incorrect (not include in source list)'.format (WO = , name = str(element)))

			for line_l in list_sep_2:
				
				if (line_l['op_nme']).lower() == element.lower() :
					
					pmt_fnl = pmt_fnl + line_l['pmt_cnt']
					
					op_lst_line = {}
					op_lst_line['se_rfr'] = line_l['se_rfr']
					op_lst_line['cn_amt'] = line_l['cn_amt']
					op_lst_line['op_lde'] = line_l['op_lde']
					op_lst_line['pmt_cnt'] = line_l['pmt_cnt']
					
					line_fnl['op_lst'].append(op_lst_line)
					
			line_fnl['pmt_fnl']= pmt_fnl
			
			list_final_by_nme.append(line_fnl)
		
		#for line in list_final_by_nme:
			#print  (line, '\n')
		
		list_final_by_nme.sort(key = lambda item: (item['op_nme'])) #sort by 'op_nme'	
		
		return list_final_by_nme

	except Exception as e:
		elog('list_sep_by_nme','{d}_{type}\n{trb1}\n{trb2} '.format(d = datetime.datetime.now(),
																	 type = str(type(e)),
																	 trb1 = traceback.format_tb(sys.exc_info()[2])[0],
																	 trb2 = str(sys.exc_info()[1])))
		return []


def list_sep_by_opt(list_payment):
	
	try:
		
		list_fnl_by_opt = []
		
		for line in list_payment:
			#print (line)
			nme_dict= []
			new_line = line
			new_line['op_nme']= str(new_line['op_nme'])
			new_line['op_nme']= new_line['op_nme'].split(',')		
			
			for name in new_line['op_nme']:
				
				if 		'FSR'.lower() 	in name.lower() :	new_name = name.replace('FSR ','').strip(' ')
				elif 	'SC'.lower() 	in name.lower() :	new_name = name.replace('SC ','').strip(' ')
				else									:	new_name = name.strip(' ')
				
				nme_dict.append(new_name)
			
			new_line['op_nme'] = nme_dict
			lnth = len(new_line['op_nme'])


			if lnth == 1:
				if float(new_line['cn_amt'])*0.01 < max_payment_lne['one_max_pmnt']:
					new_line['ttl_pmt'] = float(new_line['cn_amt'])*0.01
					new_line['one_pmt'] = float(new_line['cn_amt'])*0.01
				else:
					new_line['ttl_pmt'] = max_payment_lne['one_max_pmnt']
					new_line['one_pmt'] = max_payment_lne['one_max_pmnt']
			else:
				if float(new_line['cn_amt'])*0.015 < max_payment_lne['all_max_pmnt']:
					new_line['ttl_pmt'] = float(new_line['cn_amt'])*0.015
					new_line['one_pmt'] = (float(new_line['cn_amt'])*0.015)/lnth
				else:
					new_line['ttl_pmt'] = max_payment_lne['all_max_pmnt']
					new_line['one_pmt'] = max_payment_lne['all_max_pmnt']/lnth
					
			#print (new_line)
			
			list_fnl_by_opt.append(new_line)
			
		return list_fnl_by_opt
	
	except Exception as e:
		elog('list_sep_by_opt','{d}_{type}\n{trb1}\n{trb2} '.format(d = datetime.datetime.now(),
																	 type = str(type(e)),
																	 trb1 = traceback.format_tb(sys.exc_info()[2])[0],
																	 trb2 = str(sys.exc_info()[1])))
		return []

def make_excel(month, year, list_final_by_nme, list_final_by_opt ,filename):

	try:
		
		wb = load_workbook(filename = filename)
		ws_b = wb[u'{x}'.format(x = find_corect_list (wb))]
		
		ws_a = wb.create_sheet("Results by Name")
		ws_c = wb.create_sheet("Results by Opportunity")
		
		ws_a.sheet_view.showGridLines = False
		ws_c.sheet_view.showGridLines = False
		
		#ws_a.protection.set_password(value='FILFIL1234', already_hashed=False)
		#ws_c.protection.set_password(value='FILFIL1234', already_hashed=False)
		
		#   C A L C U L A T I O N    R E S U L T S    B Y    N A M E

		for col_index in range (1, 10):
			for row_index in range (1, 2500):
				ws_a.cell(row=row_index, column=col_index).value = ''	
		
		ws_a.column_dimensions['A'].width = 3
		ws_a.column_dimensions['B'].width = 5
		ws_a.column_dimensions['C'].width = 25
		ws_a.column_dimensions['D'].width = 20
		ws_a.column_dimensions['E'].width = 20
		ws_a.column_dimensions['F'].width = 30
		
		silver = 'FFEEEEEE'
		
		ws_a.merge_cells('B2:D2')
		ws_a.merge_cells('B3:D3')
		ws_a.merge_cells('E2:F2')
		ws_a.merge_cells('E3:F3')
		ws_a.merge_cells('B5:F5')
		
		ws_a['B2'].value = u'On the Period:'
		ws_a['B3'].value = u'Calculation date&time:'
		
		ws_a['E2'].value = date_start_month.strftime("%B %Y")
		ws_a['E3'].value = datetime.datetime.now().strftime("%d.%m.%Y  %H:%M")
		
		ws_a['B2'].alignment = Alignment(horizontal='right', vertical='center')
		ws_a['B3'].alignment = Alignment(horizontal='right', vertical='center')
		ws_a['E2'].alignment = Alignment(horizontal='right', vertical='center')
		ws_a['E3'].alignment = Alignment(horizontal='right', vertical='center')
		
		openpyxl_border_range(ws_a, 2, 2, 6, 3)
		group_fill_color(ws_a, 2, 2, 6, 3, silver)
		
		ttl_row_cur = 5
		last_row=ttl_row_cur
		
		# T O T A L    R E S U L T

		ws_a['B5'].value = u'Total Result'
		
		last_row = last_row + 1
		
		for line_n in list_final_by_nme :
			#print (line_n)
			for row_index in range(last_row , last_row + 1):
			
				ws_a.merge_cells('B{lst_r}:C{lst_r}'.format(lst_r=last_row))
				ws_a.cell(row=row_index, column=2).value = line_n['ru_nme']
				ws_a.cell(row=row_index, column=4).value= (line_n['nme']).title()
				ws_a.cell(row=row_index, column=5).value = line_n['lst_nme'].title()
				
				ws_a.cell(row=row_index, column=6).value = "= ROUND({vle} , 0)".format(vle = line_n['pmt_fnl'])
				ws_a.cell(row=row_index, column=6).data_type = 'f'	
				ws_a.cell(row=row_index, column=6).number_format = u'#,##0.00"р."'
				
				#Number of Opti:
				#ws_a.cell(row=row_index, column=7).value = len(line_n['op_lst'])
				
				last_row=last_row + 1
				
		ws_a['B5'].alignment = Alignment(horizontal='center', vertical='center')
		
		for row_index in range(ttl_row_cur + 1, last_row):
			ws_a.cell(row=row_index, column=2).alignment = Alignment(horizontal='left', vertical='center')
			ws_a.cell(row=row_index, column=4).alignment = Alignment(horizontal='left', vertical='center')
			ws_a.cell(row=row_index, column=5).alignment = Alignment(horizontal='left', vertical='center')
			ws_a.cell(row=row_index, column=6).alignment = Alignment(horizontal='left', vertical='center')
		
		openpyxl_border_range(ws_a, 2, ttl_row_cur + 1, 6, last_row - 1)
		group_fill_color(ws_a, 2, ttl_row_cur + 1, 6, last_row - 1, silver)
		
		last_row=last_row + 1
		
		# D E S C R I P T I O N
		for line_n in list_final_by_nme :
			
			ws_a.merge_cells('B{lst_r}:F{lst_r}'.format(lst_r = last_row))
			ws_a.cell(row=last_row, column=2).value = line_n['op_nme'].title()
			ws_a.cell(row=last_row, column=2).alignment= Alignment(horizontal='center', vertical='center')
			
			last_row=last_row + 1
			local_last_row=0
			index = 1
			dsc_row_cur = last_row
			
			for line_o in line_n['op_lst'] :
				
				for row_index in range(last_row, last_row + 4): 
					ws_a.merge_cells('C{row_index}:D{row_index}'.format(row_index = row_index))
					ws_a.merge_cells('E{row_index}:F{row_index}'.format(row_index = row_index))
									
				ws_a.merge_cells('B{frs_r}:B{lst_r}'.format(frs_r = last_row, lst_r = last_row + 3))
				ws_a.cell(row=last_row, column=2).value = index
				ws_a.cell(row=last_row, column=3).value = u'Opportunity Number'
				ws_a.cell(row=last_row + 1, column=3).value = u'Opportunity Leader'
				ws_a.cell(row=last_row + 2, column=3).value = u'Opportunity Amount'
				ws_a.cell(row=last_row + 3, column=3).value = u'Payment'
				
				ws_a.cell(row=last_row, column=5).value = line_o['se_rfr']
				ws_a.cell(row=last_row + 1, column=5).value = line_o['op_lde'].title()
				ws_a.cell(row=last_row + 2, column=5).value= line_o['cn_amt']
				ws_a.cell(row=last_row + 2, column=5).number_format = u'#,##0.00"р."'
				ws_a.cell(row=last_row + 3, column=5).value = line_o['pmt_cnt']
				ws_a.cell(row=last_row + 3, column=5).number_format = u'#,##0.00"р."'
				
				ws_a.cell(row=last_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
				
				for row_index in range (last_row, last_row + 4): 
					ws_a.cell(row=row_index, column=3).alignment = Alignment(horizontal='left', vertical='center')
					ws_a.cell(row=row_index, column=5).alignment = Alignment(horizontal='left', vertical='center')
				
				last_row=last_row + 5
				index	 = index + 1
			
			ws_a.merge_cells('C{row_index}:D{row_index}'.format(row_index = last_row + 1))
			ws_a.merge_cells('E{row_index}:F{row_index}'.format(row_index = last_row + 1))
			
			ws_a.cell(row=last_row, column=3).value = u'Total'
			ws_a.cell(row=last_row, column=5).value = "= ROUND({vle} , 0) ".format(vle = line_n['pmt_fnl'])
			ws_a.cell(row=last_row, column=5).number_format = u'#,##0.00"р."'
			
			ws_a.cell(row=last_row, column=3).alignment = Alignment(horizontal='left', vertical='center')
			ws_a.cell(row=last_row, column=5).alignment = Alignment(horizontal='left', vertical='center')
			
			last_row=last_row + 2	
				
			openpyxl_border_range(ws_a, 2, dsc_row_cur , 6, last_row - 2)
			group_fill_color(ws_a, 2, dsc_row_cur , 6, last_row - 2, silver)		
		
		
		#   C A L C U L A T I O N    R E S U L T S    B Y    O P P O R T U N I T Y
		for col_index in range (1, 10):
			for row_index in range (1, 1000):
				ws_c.cell(row=row_index, column=col_index).value = ''	
		
		ws_c.column_dimensions['A'].width = 3
		ws_c.column_dimensions['B'].width = 50
		ws_c.column_dimensions['C'].width = 50
		
		silver = 'FFEEEEEE'
		
		ws_c['B2'].value = u'On the Period:'
		ws_c['B3'].value = u'Calculation date&time:'
		
		ws_c['C2'].value = date_start_month.strftime("%B %Y")
		ws_c['C3'].value = datetime.datetime.now().strftime("%d.%m.%Y  %H:%M")
		
		ws_c['B2'].alignment = Alignment(horizontal='right', vertical='center')
		ws_c['B3'].alignment = Alignment(horizontal='right', vertical='center')
		ws_c['C2'].alignment = Alignment(horizontal='right', vertical='center')
		ws_c['C3'].alignment = Alignment(horizontal='right', vertical='center')
		
		openpyxl_border_range(ws_c, 2, 2, 3, 3)
		group_fill_color(ws_c, 2, 2, 3, 3, silver)
		
		ttl_row_cur = 5
		last_row=ttl_row_cur
		
		# T O T A L    R E S U L T
		ws_c.merge_cells('B5:C5')
		ws_c['B5'].value = u'Total Result'
		
		last_row=last_row + 1
		
		for line_n in list_final_by_opt :
			
			for row_index in range(last_row , last_row + 1):
			
				ws_c.cell(row=row_index, column=2).value = line_n['se_rfr']
				ws_c.cell(row=row_index, column=3).value = line_n['ttl_pmt']
				ws_c.cell(row=row_index, column=3).number_format = u'#,##0.00"р."'
				
				last_row=last_row + 1
				
		ws_c['B5'].alignment = Alignment(horizontal='center', vertical='center')
		
		for row_index in range(ttl_row_cur + 1, last_row):
			ws_c.cell(row=row_index, column=2).alignment = Alignment(horizontal='left', vertical='center')
			ws_c.cell(row=row_index, column=3).alignment = Alignment(horizontal='left', vertical='center')
		
		openpyxl_border_range(ws_c, 2, ttl_row_cur + 1, 3, last_row - 1)
		group_fill_color(ws_c, 2, ttl_row_cur + 1, 3, last_row - 1, silver)
		
		last_row=last_row + 1
		
		# D E S C R I P T I O N 
		for line_n in list_final_by_opt :
			
			ws_c.merge_cells('B{lst_r}:C{lst_r}'.format(lst_r=last_row))
			ws_c.cell(row=last_row, column=2).value = line_n['se_rfr']
			ws_c.cell(row=last_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
			
			last_row=last_row + 1
			local_last_row=0
			
			dsc_row_cur	 = last_row
			
			ws_c.cell(row=last_row, column=2).value = u'Opportunity Leader'
			ws_c.cell(row=last_row+1, column=2).value = u'Opportunity Amount'
			
			ws_c.cell(row=last_row, column=3).value = line_n['op_lde'].title()
			ws_c.cell(row=last_row+1, column=3).value= line_n['cn_amt']
			ws_c.cell(row=last_row+1, column=3).number_format = u'#,##0.00"р."'
			ws_c.cell(row=last_row+1, column=3).alignment = Alignment(horizontal='left', vertical='center')
			last_row= last_row+3
			
			for line_o in line_n['op_nme'] :
				
				ws_c.cell(row=last_row, column=2).value = line_o.title()
				ws_c.cell(row=last_row, column=3).value = line_n['one_pmt']
				ws_c.cell(row=last_row, column=3).number_format = u'#,##0.00"р."'
				ws_c.cell(row=last_row, column=3).alignment = Alignment(horizontal='left', vertical='center')
				
				last_row=last_row + 1
			
			last_row=last_row + 1
			
			ws_c.cell(row=last_row, column=2).value = u'Total'
			ws_c.cell(row=last_row, column=3).value = line_n['ttl_pmt']
			ws_c.cell(row=last_row, column=3).number_format = u'#,##0.00"р."'
			
			ws_c.cell(row=last_row, column=2).alignment = Alignment(horizontal='left', vertical='center')
			ws_c.cell(row=last_row, column=3).alignment = Alignment(horizontal='left', vertical='center')
			
			last_row=last_row + 2	
				
			openpyxl_border_range(ws_c, 2, dsc_row_cur , 3, last_row - 2)
			group_fill_color(ws_c, 2, dsc_row_cur , 3, last_row - 2, silver)		
		
		
		now = datetime.datetime.now()
		date = '{y}.{m}.{d}'.format(y=now.year, m=now.month, d=now.day)
		time = '{h}.{m}.{s}'.format(h=now.hour, m=now.minute, s=now.second)
		wb.save('{path}results\\Result_{month}_{year}_{date}_{time}.xlsx'.format(path=path,
																				  date=date,
																				  time=time,
																				  month=month,
																				  year=year
																				 )
				)
	
	except Exception as e:
		elog('make_excel','{d}_{type}\n{trb1}\n{trb2} '.format(d = datetime.datetime.now(), type = str(type(e)), trb1 = traceback.format_tb(sys.exc_info()[2])[0], trb2 = str(sys.exc_info()[1])))
		return[]

def find_corect_list(wb):
	
	#find correct list and initiate self.title
	# wb - loaded workbook
	title = ''
	try:
		str_check = 'report'
			
		for title in wb:
			if str_check in str(title): #check if including 'report' in name of Worksheet
				
				# cuting name of worksheet
				title = str(title)
				title = title.replace('<Worksheet "', '')
				title = title.replace('">', '')
		
		title = title	
		
	except	Exception as e:
		elog(
			'find_corect_list','{d}_{type}\n{trb1}\n{trb2} '.format(
				d=datetime.datetime.now(), 
				type=str(type(e)), 
				trb1=traceback.format_tb(sys.exc_info()[2])[0], 
				trb2=str(sys.exc_info()[1])
			)
		)
		return[]
	
	return title

def upload_hr_file(list_fnl):

	try:
		
		now = datetime.datetime.now()
		date_now = '{d}.{m}.{y}'.format(y = now.year, m = now.month, d = now.day)
		time_now = '{h}.{m}.{s}'.format(h = now.hour, m = now.minute, s = now.second)
		date_start= date_start_month.strftime("%d.%m.%Y")
		date_end= date_end_month.strftime("%d.%m.%Y")
		wb_hr = load_workbook(filename = '{path}templates\\Template_cis_hr.xlsm'.format(path = path), keep_vba = True)
		ws_a = wb_hr['HR request - mass change']
		
		nme_cnt = len(list_fnl)

		row_index = 11
		
		for line in list_fnl:
			ws_a['%s%s'% ('C', row_index)].value = line['ru_nme']
			ws_a['%s%s'% ('F', row_index)].value = line['nme'].title()
			ws_a['%s%s'%('G', row_index)].value = line['lst_nme'].title()
			ws_a['%s%s'%('CG', row_index)].value = "= ROUND({vle} , 0) ".format(vle = line['pmt_fnl'])
			ws_a['%s%s'%('CH', row_index)].value = u'Разовая выплата за участие в программе «Lead generation»'
			ws_a['%s%s'%('CI', row_index)].value = date_start
			ws_a['%s%s'%('CJ', row_index)].value = date_end
		
			row_index = row_index + 1
		
		wb_hr.save('{path}hr_results\\CIS_HR_request_{month}_{year}_{date}_{time}.xlsm'.format(path=path,
																								date=date_now,
																								time=time_now,
																								month=month,
																								year=year))
	
	except Exception as e:
		elog('upload_hr_file','{d}_{type}\n{trb1}\n{trb2} '.format(d=datetime.datetime.now(),
																	type=str(type(e)),
																	trb1=traceback.format_tb(sys.exc_info()[2])[0],
																	trb2=str(sys.exc_info()[1])
																   )
			 )
		return[]	
		
month= 6
year= 2018

max_payment_lne = {}

for line in max_payment_lst:
	if line['yer'] == year:
		max_payment_lne = line

path = 'path'
filename = ''.format(path=path,month=month, year=year)

date_start_month= datetime.date(year, month	, 1)
if month == 12:
	date_end_month = datetime.date(year + 1, 1, 1) - timedelta(days = 1)
else:
	date_end_month = datetime.date(year, month + 1, 1) - timedelta(days = 1)

list_payment_1 = load_bfs(filename)
list_payment_2 = load_bfs(filename)
list_final_by_opt = list_sep_by_opt(list_payment_1) 
list_final_by_nme = list_sep_by_nme(list_payment_2) 

make_excel(month, year, list_final_by_nme, list_final_by_opt, filename)
upload_hr_file(list_final_by_nme)
